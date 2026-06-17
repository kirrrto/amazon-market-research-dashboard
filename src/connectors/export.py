from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.i18n import column_label, normalize_language, sheet_label
from src.requirements import (
    build_decision_summary,
    build_requirement_draft,
    build_supplier_follow_up_questions,
)
from src.specs import (
    build_coverage_summary,
    build_gap_analysis,
    build_specification_matrix,
    normalize_raw_specifications,
)

from .models import FetchResult


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        values = [str(cell.value or "") for cell in column_cells[:100]]
        width = min(max(max(map(len, values), default=8) + 2, 10), 60)
        worksheet.column_dimensions[letter].width = width


def _format_percent_columns(worksheet, headers_to_format: set[str]) -> None:
    headers = {
        str(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    for header in headers_to_format:
        column_index = headers.get(header)
        if column_index is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_index).number_format = "0.00%"


def _localized_frame(frame: pd.DataFrame, language: str) -> pd.DataFrame:
    return frame.rename(columns={column: column_label(str(column), language) for column in frame.columns})


def normalized_specs_frame(result: FetchResult, language: str = "en") -> pd.DataFrame:
    """Build a normalized specification frame from raw connector specifications."""

    raw_frame = result.raw_specs_frame
    if raw_frame.empty:
        columns = [
            "source_url",
            "standard_field",
            "standard_label",
            "raw_spec_name",
            "raw_spec_value",
            "normalized_value",
            "confidence",
            "category",
            "parser",
            "fetched_at",
        ]
        return pd.DataFrame(columns=columns)

    rows = normalize_raw_specifications(
        raw_frame.to_dict("records"),
        language=language,
    )
    columns = [
        "source_url",
        "standard_field",
        "standard_label",
        "raw_spec_name",
        "raw_spec_value",
        "normalized_value",
        "confidence",
        "category",
        "parser",
        "fetched_at",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def specification_matrix_frame(result: FetchResult, language: str = "en") -> pd.DataFrame:
    normalized = normalized_specs_frame(result, language=language)
    return build_specification_matrix(normalized, products=result.products_frame)


def coverage_summary_frame(result: FetchResult, language: str = "en") -> pd.DataFrame:
    normalized = normalized_specs_frame(result, language=language)
    return build_coverage_summary(
        normalized,
        products=result.products_frame,
        language=language,
    )


def gap_analysis_frame(
    result: FetchResult,
    *,
    profile: str = "generic_hardware",
    language: str = "en",
) -> pd.DataFrame:
    matrix = specification_matrix_frame(result, language=language)
    return build_gap_analysis(matrix, profile=profile, language=language)


def requirement_draft_frame(
    result: FetchResult,
    *,
    profile: str = "generic_hardware",
    language: str = "en",
) -> pd.DataFrame:
    matrix = specification_matrix_frame(result, language=language)
    gaps = gap_analysis_frame(result, profile=profile, language=language)
    return build_requirement_draft(
        matrix,
        gap_analysis=gaps,
        profile=profile,
        language=language,
    )


def supplier_follow_up_frame(
    result: FetchResult,
    *,
    profile: str = "generic_hardware",
    language: str = "en",
) -> pd.DataFrame:
    gaps = gap_analysis_frame(result, profile=profile, language=language)
    return build_supplier_follow_up_questions(gaps, language=language)


def decision_summary_frame(
    result: FetchResult,
    *,
    profile: str = "generic_hardware",
    language: str = "en",
) -> pd.DataFrame:
    gaps = gap_analysis_frame(result, profile=profile, language=language)
    return build_decision_summary(gaps, language=language)


def build_product_page_workbook(
    result: FetchResult,
    language: str = "en",
    profile: str = "generic_hardware",
) -> bytes:
    """Build a workbook for public product page connector results."""

    code = normalize_language(language)
    buffer = BytesIO()

    products = _localized_frame(result.products_frame, code)
    raw_specs = _localized_frame(result.raw_specs_frame, code)
    normalized_specs = _localized_frame(normalized_specs_frame(result, code), code)
    matrix = _localized_frame(specification_matrix_frame(result, code), code)
    coverage = _localized_frame(coverage_summary_frame(result, code), code)
    gaps = _localized_frame(gap_analysis_frame(result, profile=profile, language=code), code)
    requirement_draft = _localized_frame(
        requirement_draft_frame(result, profile=profile, language=code),
        code,
    )
    supplier_follow_up = _localized_frame(
        supplier_follow_up_frame(result, profile=profile, language=code),
        code,
    )
    decision_summary = _localized_frame(
        decision_summary_frame(result, profile=profile, language=code),
        code,
    )
    fetch_logs = _localized_frame(result.fetch_logs_frame, code)
    issues = _localized_frame(result.issues_frame, code)

    sheet_names = [
        sheet_label("products", code),
        sheet_label("raw_specifications", code),
        sheet_label("normalized_specifications", code),
        sheet_label("specification_matrix", code),
        sheet_label("coverage_summary", code),
        sheet_label("gap_analysis", code),
        sheet_label("requirement_draft", code),
        sheet_label("supplier_follow_up", code),
        sheet_label("decision_summary", code),
        sheet_label("fetch_logs", code),
        sheet_label("issues", code),
    ]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        products.to_excel(writer, sheet_name=sheet_names[0], index=False)
        raw_specs.to_excel(writer, sheet_name=sheet_names[1], index=False)
        normalized_specs.to_excel(writer, sheet_name=sheet_names[2], index=False)
        matrix.to_excel(writer, sheet_name=sheet_names[3], index=False)
        coverage.to_excel(writer, sheet_name=sheet_names[4], index=False)
        gaps.to_excel(writer, sheet_name=sheet_names[5], index=False)
        requirement_draft.to_excel(writer, sheet_name=sheet_names[6], index=False)
        supplier_follow_up.to_excel(writer, sheet_name=sheet_names[7], index=False)
        decision_summary.to_excel(writer, sheet_name=sheet_names[8], index=False)
        fetch_logs.to_excel(writer, sheet_name=sheet_names[9], index=False)
        issues.to_excel(writer, sheet_name=sheet_names[10], index=False)

        for sheet_name in sheet_names:
            _style_sheet(writer.book[sheet_name])

        _format_percent_columns(
            writer.book[sheet_names[4]],
            {column_label("coverage_rate", code), "coverage_rate"},
        )
        _format_percent_columns(
            writer.book[sheet_names[5]],
            {column_label("completion_rate", code), "completion_rate"},
        )
        _format_percent_columns(
            writer.book[sheet_names[8]],
            {column_label("completion_rate", code), "completion_rate"},
        )

    buffer.seek(0)
    return buffer.getvalue()
