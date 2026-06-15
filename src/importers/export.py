from __future__ import annotations

from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.finance import ProfitAssumptions, add_profit_estimates
from src.i18n import column_label, normalize_language, sheet_label

from .models import ImportResult


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(worksheet, freeze_panes: str = "A2") -> None:
    worksheet.freeze_panes = freeze_panes
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        values = [str(cell.value or "") for cell in column_cells[:100]]
        width = min(max(max(map(len, values), default=8) + 2, 10), 40)
        worksheet.column_dimensions[letter].width = width


def _format_products_sheet(worksheet, language: str = "en") -> None:
    headers = {
        str(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    currency_columns = {
        "price",
        "estimated_monthly_revenue",
        "estimated_unit_gross_profit",
        "estimated_unit_net_profit",
        "break_even_price",
        "estimated_monthly_net_profit",
    }
    percentage_columns = {"estimated_net_margin"}

    for header in currency_columns:
        localized_header = column_label(header, language)
        column_index = headers.get(localized_header) or headers.get(header)
        if column_index is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_index).number_format = '$#,##0.00'

    for header in percentage_columns:
        localized_header = column_label(header, language)
        column_index = headers.get(localized_header) or headers.get(header)
        if column_index is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_index).number_format = '0.00%'


def _localized_frame(frame: pd.DataFrame, language: str) -> pd.DataFrame:
    return frame.rename(columns={column: column_label(str(column), language) for column in frame.columns})


def prepare_export_result(
    result: ImportResult,
    assumptions: ProfitAssumptions,
) -> ImportResult:
    """Return an export-ready result using the current profit assumptions."""

    return ImportResult(
        products=add_profit_estimates(result.products, assumptions),
        issues=result.issues,
        report=result.report,
    )


def build_normalized_workbook(result: ImportResult, language: str = "en") -> bytes:
    code = normalize_language(language)
    buffer = BytesIO()
    issues = _localized_frame(result.issues_frame, code)
    products = _localized_frame(result.products, code)

    summary_headers = (
        ["Item", "Value"]
        if code == "en"
        else ["项目", "值"]
    )
    mapping_headers = (
        ["Standard field", "Source column"]
        if code == "en"
        else ["标准字段", "源字段"]
    )

    summary = pd.DataFrame(
        result.report.summary_rows(),
        columns=summary_headers,
    )
    mapping = pd.DataFrame(
        sorted(result.report.mapping.items()),
        columns=mapping_headers,
    )

    products_sheet_name = sheet_label("products", code)
    issues_sheet_name = sheet_label("issues", code)
    report_sheet_name = sheet_label("import_report", code)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        products.to_excel(writer, sheet_name=products_sheet_name, index=False)
        issues.to_excel(writer, sheet_name=issues_sheet_name, index=False)
        summary.to_excel(
            writer,
            sheet_name=report_sheet_name,
            index=False,
            startrow=0,
        )
        mapping.to_excel(
            writer,
            sheet_name=report_sheet_name,
            index=False,
            startrow=len(summary) + 3,
        )

        products_sheet = writer.book[products_sheet_name]
        _style_sheet(products_sheet)
        _format_products_sheet(products_sheet, code)
        _style_sheet(writer.book[issues_sheet_name])
        report_sheet = writer.book[report_sheet_name]
        _style_sheet(report_sheet)
        mapping_header_row = len(summary) + 4
        for cell in report_sheet[mapping_header_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    buffer.seek(0)
    return buffer.getvalue()
