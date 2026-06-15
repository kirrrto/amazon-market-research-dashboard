from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.i18n import normalize_language, sheet_label

from .schemas import TemplateDefinition, TemplateField, get_template_definition


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="E2F0D9")


def _field_label(field: TemplateField, language: str) -> str:
    return field.label_zh if language == "zh-CN" else field.label_en


def _field_description(field: TemplateField, language: str) -> str:
    return field.description_zh if language == "zh-CN" else field.description_en


def _style_header(sheet: Worksheet, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autofit(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        values = [str(cell.value or "") for cell in column_cells[:100]]
        width = min(max(max(map(len, values), default=8) + 2, 12), 44)
        sheet.column_dimensions[letter].width = width


def _create_data_sheet(workbook: Workbook, definition: TemplateDefinition, language: str) -> None:
    sheet = workbook.active
    sheet.title = sheet_label(definition.sheet_key, language)

    headers = [field.name for field in definition.fields]
    labels = [_field_label(field, language) for field in definition.fields]
    examples = [field.example for field in definition.fields]

    sheet.append(headers)
    sheet.append(labels)
    sheet.append(examples)
    _style_header(sheet, 1)

    for cell in sheet[2]:
        cell.fill = NOTE_FILL

    sheet.freeze_panes = "A2"
    _autofit(sheet)


def _create_dictionary_sheet(workbook: Workbook, definition: TemplateDefinition, language: str) -> None:
    sheet = workbook.create_sheet(sheet_label("data_dictionary", language))
    headers = {
        "en": ["Field", "Label", "Required", "Description", "Example"],
        "zh-CN": ["字段", "显示名称", "是否必填", "说明", "示例"],
    }[language]
    sheet.append(headers)
    _style_header(sheet, 1)

    for field in definition.fields:
        sheet.append(
            [
                field.name,
                _field_label(field, language),
                "Yes" if field.required and language == "en" else "是" if field.required else "No" if language == "en" else "否",
                _field_description(field, language),
                field.example,
            ]
        )

    sheet.freeze_panes = "A2"
    _autofit(sheet)


def _create_example_sheet(workbook: Workbook, definition: TemplateDefinition, language: str) -> None:
    sheet = workbook.create_sheet(sheet_label("example", language))
    sheet.append([field.name for field in definition.fields])
    sheet.append([field.example for field in definition.fields])
    _style_header(sheet, 1)
    sheet.freeze_panes = "A2"
    _autofit(sheet)


def build_template_workbook(template_id: str, language: str = "en") -> bytes:
    code = normalize_language(language)
    definition = get_template_definition(template_id)

    workbook = Workbook()
    _create_data_sheet(workbook, definition, code)
    _create_dictionary_sheet(workbook, definition, code)
    _create_example_sheet(workbook, definition, code)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
