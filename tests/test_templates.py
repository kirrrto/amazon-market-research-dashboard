from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.templates import (
    build_template_workbook,
    get_template_definition,
    list_template_definitions,
)


def workbook_sheet_names(template_id: str, language: str = "en") -> list[str]:
    workbook = load_workbook(BytesIO(build_template_workbook(template_id, language)))
    return workbook.sheetnames


def test_lists_expected_templates():
    template_ids = {item.template_id for item in list_template_definitions()}
    assert {
        "market_research",
        "supplier_quote",
        "product_url_import",
        "security_camera_specs",
        "vehicle_camera_specs",
    }.issubset(template_ids)


def test_unknown_template_id_raises_key_error():
    with pytest.raises(KeyError):
        get_template_definition("missing_template")


def test_market_research_template_has_required_fields():
    definition = get_template_definition("market_research")
    required_fields = {field.name for field in definition.fields if field.required}
    assert {
        "asin",
        "title",
        "brand",
        "price",
        "rating",
        "reviews",
        "monthly_sales",
        "category",
    }.issubset(required_fields)


def test_product_url_template_contains_source_url():
    definition = get_template_definition("product_url_import")
    fields = {field.name for field in definition.fields}
    assert "source_url" in fields


def test_security_camera_template_contains_industry_fields():
    definition = get_template_definition("security_camera_specs")
    fields = {field.name for field in definition.fields}
    assert {"resolution", "waterproof_rating", "battery_capacity_mah", "supports_rtsp"}.issubset(fields)


def test_vehicle_camera_template_contains_industry_fields():
    definition = get_template_definition("vehicle_camera_specs")
    fields = {field.name for field in definition.fields}
    assert {"camera_resolution", "power_input", "waterproof_rating", "carplay_interference_risk"}.issubset(fields)


def test_english_workbook_contains_expected_sheets():
    assert workbook_sheet_names("market_research", "en") == [
        "Products",
        "Data Dictionary",
        "Example",
    ]


def test_chinese_workbook_contains_expected_sheets():
    assert workbook_sheet_names("product_url_import", "zh-CN") == [
        "产品页面链接",
        "数据字典",
        "示例",
    ]


def test_data_sheet_contains_field_names_and_labels():
    workbook = load_workbook(BytesIO(build_template_workbook("market_research", "zh-CN")))
    sheet = workbook["产品数据"]
    assert sheet["A1"].value == "asin"
    assert sheet["A2"].value == "ASIN"
    assert sheet["B2"].value == "标题"


def test_dictionary_sheet_contains_descriptions():
    workbook = load_workbook(BytesIO(build_template_workbook("supplier_quote", "zh-CN")))
    sheet = workbook["数据字典"]
    values = [cell.value for cell in sheet["A"]]
    assert "supplier_name" in values
    assert "firmware_ownership" in values


def test_all_templates_build_in_both_languages():
    for definition in list_template_definitions():
        for language in ("en", "zh-CN"):
            workbook = load_workbook(BytesIO(build_template_workbook(definition.template_id, language)))
            assert len(workbook.sheetnames) == 3
